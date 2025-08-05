#!/usr/bin/env python3
"""
Create sample Excel sheets based on the Stark audit data
This script recreates the structure of the original audit data request sheets
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_primary_audit_sheet():
    """Create the primary audit questions sheet"""
    
    # Sample data based on the Stark audit structure
    primary_questions = [
        {
            'Question ID': 'Q1',
            'Question': 'Describe the database management controls and procedures in place for ensuring data integrity and backup processes.',
            'Category': 'Database Management',
            'Priority': 'High',
            'Expected Response Type': 'Detailed description'
        },
        {
            'Question ID': 'Q2', 
            'Question': 'What change management procedures are followed for system modifications and updates?',
            'Category': 'Change Management',
            'Priority': 'High',
            'Expected Response Type': 'Process documentation'
        },
        {
            'Question ID': 'Q3',
            'Question': 'Detail the security controls implemented for user access management and authentication.',
            'Category': 'Security Controls',
            'Priority': 'Critical',
            'Expected Response Type': 'Security framework'
        },
        {
            'Question ID': 'Q4',
            'Question': 'Describe the computer operations monitoring and incident response procedures.',
            'Category': 'Computer Operations',
            'Priority': 'High',
            'Expected Response Type': 'Operational procedures'
        },
        {
            'Question ID': 'Q5',
            'Question': 'What backup and recovery procedures are in place for critical systems and data?',
            'Category': 'Backup & Recovery',
            'Priority': 'Critical',
            'Expected Response Type': 'Recovery plans'
        },
        {
            'Question ID': 'Q6',
            'Question': 'Provide details on general IT governance and compliance framework.',
            'Category': 'General',
            'Priority': 'Medium',
            'Expected Response Type': 'Framework documentation'
        },
        {
            'Question ID': 'Q7',
            'Question': 'How are database performance metrics monitored and reported?',
            'Category': 'Database Management',
            'Priority': 'Medium',
            'Expected Response Type': 'Monitoring reports'
        },
        {
            'Question ID': 'Q8',
            'Question': 'What approval workflows exist for emergency changes to production systems?',
            'Category': 'Change Management',
            'Priority': 'High',
            'Expected Response Type': 'Workflow documentation'
        },
        {
            'Question ID': 'Q9',
            'Question': 'Detail the network security controls and firewall configurations.',
            'Category': 'Security Controls',
            'Priority': 'High',
            'Expected Response Type': 'Network diagrams'
        },
        {
            'Question ID': 'Q10',
            'Question': 'Describe the system maintenance schedules and downtime procedures.',
            'Category': 'Computer Operations',
            'Priority': 'Medium',
            'Expected Response Type': 'Maintenance schedules'
        },
        {
            'Question ID': 'Q11',
            'Question': 'What are the data retention policies for backup storage?',
            'Category': 'Backup & Recovery',
            'Priority': 'Medium',
            'Expected Response Type': 'Policy documents'
        },
        {
            'Question ID': 'Q12',
            'Question': 'How is compliance with regulatory requirements monitored and reported?',
            'Category': 'General',
            'Priority': 'High',
            'Expected Response Type': 'Compliance reports'
        },
        {
            'Question ID': 'Q13',
            'Question': 'Detail the database encryption and data protection measures.',
            'Category': 'Database Management',
            'Priority': 'Critical',
            'Expected Response Type': 'Encryption standards'
        },
        {
            'Question ID': 'Q14',
            'Question': 'What testing procedures are followed for system changes before production deployment?',
            'Category': 'Change Management',
            'Priority': 'High',
            'Expected Response Type': 'Testing protocols'
        },
        {
            'Question ID': 'Q15',
            'Question': 'Describe the vulnerability management and patch deployment process.',
            'Category': 'Security Controls',
            'Priority': 'Critical',
            'Expected Response Type': 'Security procedures'
        },
        {
            'Question ID': 'Q16',
            'Question': 'How are system performance issues identified and resolved?',
            'Category': 'Computer Operations',
            'Priority': 'Medium',
            'Expected Response Type': 'Performance metrics'
        },
        {
            'Question ID': 'Q17',
            'Question': 'What disaster recovery testing is performed and how frequently?',
            'Category': 'Backup & Recovery',
            'Priority': 'Critical',
            'Expected Response Type': 'DR test results'
        },
        {
            'Question ID': 'Q18',
            'Question': 'Detail the vendor management and third-party risk assessment procedures.',
            'Category': 'General',
            'Priority': 'Medium',
            'Expected Response Type': 'Risk assessments'
        },
        {
            'Question ID': 'Q19',
            'Question': 'How are database user privileges managed and reviewed?',
            'Category': 'Database Management',
            'Priority': 'High',
            'Expected Response Type': 'Access reviews'
        },
        {
            'Question ID': 'Q20',
            'Question': 'What documentation standards are maintained for system changes?',
            'Category': 'Change Management',
            'Priority': 'Medium',
            'Expected Response Type': 'Documentation standards'
        }
    ]
    
    df = pd.DataFrame(primary_questions)
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Primary Audit Questions"
    
    # Add header row with formatting
    headers = ['Question ID', 'Question', 'Category', 'Priority', 'Expected Response Type']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == 1:  # Question ID column
                cell.font = Font(bold=True)
            elif c_idx == 4:  # Priority column
                if value == 'Critical':
                    cell.font = Font(color='FF0000', bold=True)
                elif value == 'High':
                    cell.font = Font(color='FF6600', bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Question ID
    ws.column_dimensions['B'].width = 80  # Question
    ws.column_dimensions['C'].width = 20  # Category
    ws.column_dimensions['D'].width = 12  # Priority
    ws.column_dimensions['E'].width = 25  # Expected Response Type
    
    return wb

def create_followup_audit_sheet():
    """Create the followup audit questions sheet"""
    
    # Sample followup questions based on the Stark audit structure
    followup_questions = [
        {
            'Question ID': 'FQ1',
            'Question': 'Based on the initial security controls assessment, provide detailed penetration testing results and remediation status.',
            'Category': 'Follow-up Security',
            'Related to Primary': 'Q3, Q9, Q15',
            'Expected Response Type': 'Security test results'
        },
        {
            'Question ID': 'FQ2',
            'Question': 'Following the change management review, detail any exceptions or deviations from standard procedures in the last quarter.',
            'Category': 'Follow-up Operations', 
            'Related to Primary': 'Q2, Q8, Q14',
            'Expected Response Type': 'Exception reports'
        },
        {
            'Question ID': 'FQ3',
            'Question': 'Provide updated disaster recovery test results and any improvements made since the last assessment.',
            'Category': 'Follow-up Operations',
            'Related to Primary': 'Q5, Q17',
            'Expected Response Type': 'Updated DR plans'
        },
        {
            'Question ID': 'FQ4',
            'Question': 'Detail any security incidents that occurred since the primary assessment and response actions taken.',
            'Category': 'Follow-up Security',
            'Related to Primary': 'Q3, Q4',
            'Expected Response Type': 'Incident reports'
        },
        {
            'Question ID': 'FQ5',
            'Question': 'Provide compliance audit findings and corrective actions implemented for regulatory requirements.',
            'Category': 'Follow-up General',
            'Related to Primary': 'Q6, Q12, Q18',
            'Expected Response Type': 'Compliance updates'
        },
        {
            'Question ID': 'FQ6',
            'Question': 'Following database management review, provide current performance metrics and optimization measures.',
            'Category': 'Follow-up Operations',
            'Related to Primary': 'Q1, Q7, Q13, Q19',
            'Expected Response Type': 'Performance data'
        },
        {
            'Question ID': 'FQ7',
            'Question': 'Detail any changes to user access controls and authentication mechanisms since the primary review.',
            'Category': 'Follow-up Security',
            'Related to Primary': 'Q3, Q19',
            'Expected Response Type': 'Access control updates'
        },
        {
            'Question ID': 'FQ8',
            'Question': 'Provide status on any outstanding recommendations from the primary audit assessment.',
            'Category': 'Follow-up General',
            'Related to Primary': 'All questions',
            'Expected Response Type': 'Recommendation status'
        }
    ]
    
    df = pd.DataFrame(followup_questions)
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Followup Audit Questions"
    
    # Add header row with formatting
    headers = ['Question ID', 'Question', 'Category', 'Related to Primary', 'Expected Response Type']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows  
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == 1:  # Question ID column
                cell.font = Font(bold=True, color='8B4513')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Question ID
    ws.column_dimensions['B'].width = 80  # Question
    ws.column_dimensions['C'].width = 25  # Category
    ws.column_dimensions['D'].width = 20  # Related to Primary
    ws.column_dimensions['E'].width = 25  # Expected Response Type
    
    return wb

if __name__ == "__main__":
    # Create the sample sheets
    print("Creating sample audit data request sheets...")
    
    # Create primary sheet
    primary_wb = create_primary_audit_sheet()
    primary_filename = "SAMPLE_Primary_Audit_Questions.xlsx"
    primary_wb.save(primary_filename)
    print(f"✓ Created: {primary_filename}")
    
    # Create followup sheet
    followup_wb = create_followup_audit_sheet()
    followup_filename = "SAMPLE_Followup_Audit_Questions.xlsx"
    followup_wb.save(followup_filename)
    print(f"✓ Created: {followup_filename}")
    
    print("\nSample sheets created successfully!")
    print("These sheets demonstrate the structure used in the Stark G4 Q2CA2025 audit.")